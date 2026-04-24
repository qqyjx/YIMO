"""EAV 快速批量导入 (eav_full.py 的加速版).

优化:
 - executemany 批量 INSERT (pymysql 自动重写为单条 bulk VALUES)
 - UNIQUE_CHECKS=0 + FOREIGN_KEY_CHECKS=0 导入期禁用
 - 简化类型推断: 全部按 TEXT 存 value_text (DA 表本身就是文本为主)
 - 跳过增量/row_hash 计算 (整域重灌)
 - 批量读 Excel 到内存 (openpyxl read_only=False + data_only=True)
 - lastrowid 推导连续 entity_id 段 (避免 INSERT 后 SELECT id 回查)

预期: 100 row/s → 1500-3000 row/s, 88k 行 DA-02 从 15 分钟 → 30-60 秒.

CLI 与 eav_full.py 保持兼容:
    python eav_fast_import.py --excel <path> --data-domain <domain>
        --dataset-name <name> --host 127.0.0.1 --port 3307
        --user eav_user --password eavpass123 --db eav_db
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
import pymysql


BATCH_ENTITIES = 5000    # 一次 bulk insert 多少行 entities
BATCH_VALUES = 10000     # 一次 bulk insert 多少行 values


def normalize_attr_name(s) -> str:
    """属性名规范化: 去空格/换行, 截断到 160 字符."""
    if s is None:
        return ""
    s = str(s).strip().replace("\n", " ").replace("\r", " ")
    return s[:160] if len(s) > 160 else s


def val_to_text(v) -> str | None:
    """全部当文本存, None/NaN 转 NULL."""
    if v is None:
        return None
    if isinstance(v, float):
        if v != v:  # NaN
            return None
    s = str(v).strip()
    if s == "" or s.lower() in ("nan", "none", "null"):
        return None
    return s


def load_excel(path: Path) -> dict[str, list[dict]]:
    """读 xlsx 到 {sheet_name: [row_dict, ...]}. 返回 row 字典列表避免 pandas 依赖."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        if (ws.max_row or 0) <= 1:
            continue
        rows_iter = ws.iter_rows(values_only=True)
        # 找首个非空 header 行
        header = None
        skip = 0
        for r in rows_iter:
            skip += 1
            if r and any(c not in (None, "") for c in r):
                header = [normalize_attr_name(c) for c in r]
                break
        if header is None:
            continue
        # 去重复空列
        header = [h if h else f"col_{i}" for i, h in enumerate(header)]
        rows = []
        for r in rows_iter:
            if r is None:
                continue
            if all(c in (None, "") for c in r):
                continue
            rows.append({h: r[i] if i < len(r) else None for i, h in enumerate(header)})
        if rows:
            result[sn] = {"header": header, "rows": rows}
    wb.close()
    return result


def main() -> int:
    ap = argparse.ArgumentParser(description="EAV 快速批量导入 (bulk INSERT 版)")
    ap.add_argument("--excel", required=True)
    ap.add_argument("--dataset-name", required=True)
    ap.add_argument("--data-domain", default="default")
    ap.add_argument("--host", default="127.0.0.1")
    ap.add_argument("--port", type=int, default=3307)
    ap.add_argument("--user", default="eav_user")
    ap.add_argument("--password", default="eavpass123")
    ap.add_argument("--db", default="eav_db")
    ap.add_argument("--lifecycle-stage", default="Operation")
    args = ap.parse_args()

    path = Path(args.excel)
    if not path.is_file():
        print(f"[ERR] 文件不存在: {path}", file=sys.stderr)
        return 2

    t0 = time.time()
    print(f"[INFO] 读 Excel: {path.name}")
    sheets = load_excel(path)
    total_rows = sum(len(v["rows"]) for v in sheets.values())
    print(f"[INFO] Excel 含 {len(sheets)} sheet, 合计 {total_rows:,} 行")
    t_load = time.time() - t0

    conn = pymysql.connect(
        host=args.host, port=args.port, user=args.user,
        password=args.password, database=args.db,
        charset="utf8mb4", autocommit=False,
    )
    cur = conn.cursor()

    # 1) 导入前禁用约束 (session 级)
    cur.execute("SET UNIQUE_CHECKS=0")
    cur.execute("SET FOREIGN_KEY_CHECKS=0")
    cur.execute("SET autocommit=0")

    # 2) 创建 dataset
    cur.execute("""
        INSERT INTO eav_datasets
            (name, data_domain, source_file, lifecycle_stage, imported_at)
        VALUES (%s, %s, %s, %s, %s)
    """, (args.dataset_name, args.data_domain, str(path),
          args.lifecycle_stage, datetime.utcnow()))
    dataset_id = cur.lastrowid
    print(f"[INFO] dataset_id = {dataset_id}")

    # 3) 汇总全局 attribute 集合 (跨 sheet 合并同名)
    all_attrs = {}    # name → display_name 首次出现
    for sn, meta in sheets.items():
        for h in meta["header"]:
            if h and h not in all_attrs:
                all_attrs[h] = h
    # 加 __sheet__ 系统属性
    all_attrs["__sheet__"] = "__sheet__"

    # 4) bulk INSERT attributes
    attr_rows = [
        (dataset_id, name, name, "text", i)
        for i, name in enumerate(all_attrs.keys())
    ]
    cur.executemany("""
        INSERT INTO eav_attributes (dataset_id, name, display_name, data_type, ord_index)
        VALUES (%s, %s, %s, %s, %s)
    """, attr_rows)
    # 取回 name → id
    cur.execute("SELECT id, name FROM eav_attributes WHERE dataset_id=%s", (dataset_id,))
    attr_ids = {name: _id for _id, name in cur.fetchall()}
    print(f"[INFO] 已建 {len(attr_ids)} 属性")

    # 5) 每 sheet 走 bulk INSERT entities + values
    t_insert = time.time()
    total_ent = 0
    total_val = 0
    global_row_no = 0

    for sn, meta in sheets.items():
        rows = meta["rows"]
        header = meta["header"]
        n = len(rows)
        print(f"[INFO] sheet={sn} {n:,} 行")

        # 5a. bulk INSERT entities 分批
        ent_buffer = []
        ent_ids_this_sheet = []
        for row_data in rows:
            global_row_no += 1
            ent_buffer.append((dataset_id, global_row_no, None, "", datetime.utcnow()))
            if len(ent_buffer) >= BATCH_ENTITIES:
                cur.executemany("""
                    INSERT INTO eav_entities (dataset_id, `row_number`, external_id, row_hash, created_at)
                    VALUES (%s, %s, %s, %s, %s)
                """, ent_buffer)
                first_id = cur.lastrowid
                ent_ids_this_sheet.extend(range(first_id, first_id + len(ent_buffer)))
                ent_buffer = []
        if ent_buffer:
            cur.executemany("""
                INSERT INTO eav_entities (dataset_id, `row_number`, external_id, row_hash, created_at)
                VALUES (%s, %s, %s, %s, %s)
            """, ent_buffer)
            first_id = cur.lastrowid
            ent_ids_this_sheet.extend(range(first_id, first_id + len(ent_buffer)))
            ent_buffer = []

        assert len(ent_ids_this_sheet) == n, \
            f"entity id 段推导异常 ({len(ent_ids_this_sheet)} vs {n})"
        total_ent += n

        # 5b. bulk INSERT values 分批 (跨列拍平)
        val_buffer = []
        sheet_attr_id = attr_ids.get("__sheet__")
        for ent_id, row_data in zip(ent_ids_this_sheet, rows):
            # __sheet__ 系统属性
            if sheet_attr_id:
                val_buffer.append((ent_id, sheet_attr_id, sn, None, None, None, sn))
            # 每列一个 value
            for col in header:
                aid = attr_ids.get(col)
                if not aid:
                    continue
                v = val_to_text(row_data.get(col))
                if v is None:
                    continue   # 空值不入 (节省空间)
                val_buffer.append((ent_id, aid, v, None, None, None, v))
                if len(val_buffer) >= BATCH_VALUES:
                    cur.executemany("""
                        INSERT INTO eav_values
                            (entity_id, attribute_id, value_text, value_number, value_datetime, value_bool, raw_text)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, val_buffer)
                    total_val += len(val_buffer)
                    val_buffer = []
        if val_buffer:
            cur.executemany("""
                INSERT INTO eav_values
                    (entity_id, attribute_id, value_text, value_number, value_datetime, value_bool, raw_text)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, val_buffer)
            total_val += len(val_buffer)
        conn.commit()   # 每 sheet 结束一次提交
        print(f"[INFO]   sheet 完成, entities={n:,}")

    # 6) 恢复约束
    cur.execute("SET UNIQUE_CHECKS=1")
    cur.execute("SET FOREIGN_KEY_CHECKS=1")
    conn.commit()
    conn.close()

    t_total = time.time() - t0
    t_ins = time.time() - t_insert
    print(f"\n[DONE] {args.data_domain}: entities={total_ent:,}, values={total_val:,}")
    print(f"       Excel 读取 {t_load:.1f}s | DB 写入 {t_ins:.1f}s | 总 {t_total:.1f}s")
    if t_ins > 0:
        print(f"       吞吐: {total_ent / t_ins:,.0f} entities/s, {total_val / t_ins:,.0f} values/s")
    return 0


if __name__ == "__main__":
    sys.exit(main())
