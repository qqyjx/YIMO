"""Process a single xlsx: drop columns whose last-in-row header contains 校验."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl


def drop_verification_columns(xlsx_path: Path) -> list[tuple[str, str]]:
    removed: list[tuple[str, str]] = []
    wb = openpyxl.load_workbook(xlsx_path)
    for sn in wb.sheetnames:
        ws = wb[sn]
        max_col = ws.max_column
        if max_col == 0:
            continue
        header_row_idx = None
        for r in range(1, min(6, ws.max_row or 0) + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if any(v not in (None, "") for v in row_vals):
                header_row_idx = r
                break
        if header_row_idx is None:
            continue
        last_col = None
        last_hdr = None
        for c in range(max_col, 0, -1):
            v = ws.cell(row=header_row_idx, column=c).value
            if v not in (None, ""):
                last_col = c
                last_hdr = v
                break
        if last_col is None:
            continue
        if "校验" in str(last_hdr):
            ws.delete_cols(last_col, 1)
            removed.append((sn, str(last_hdr)))
    if removed:
        wb.save(xlsx_path)
    wb.close()
    return removed


if __name__ == "__main__":
    path = Path(sys.argv[1])
    removed = drop_verification_columns(path)
    print(f"removed={len(removed)}", flush=True)
