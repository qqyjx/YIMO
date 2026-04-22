#!/usr/bin/env python3
"""
批量导入南方电网全域三架构 Excel 到 DATA/<domain>/，并清洗：
  - 删除每个 sheet 最右侧的"校验结果"列
  - 删除无关 sheet（sheet1 / WpsReserved / 说明 / 临时-* / 导出计数_*）

用法：
  python scripts/import_all_domains.py \\
      --source-dir /tmp/yimo-new-data \\
      --data-dir DATA \\
      [--dry-run]

约束：
  - 不覆盖 DATA/shupeidian/ 与 DATA/jicai/ 下的已有文件
  - 输出命名约定：1.xlsx=业务  2.xlsx=数据  3.xlsx=应用
"""
from __future__ import annotations
import argparse
import re
import shutil
from pathlib import Path

import openpyxl

DOMAIN_MAP = {
    '纪检监察': 'jjjc',
    '党建': 'dangjian',
    '战略规划': 'zhanlve',
    '市场营销': 'yingxiao',
    '新兴业务': 'xinxing',
    '人力资源': 'hr',
    '企业架构': 'qyjg',
    '政策研究': 'zhengce',
    '安全监管': 'anquan',
    '巡视': 'xunshi',
    '工会': 'gonghui',
    '审计': 'shenji',
    '供应链': 'gongying',
    '系统运行': 'xitong',
    '办公': 'bangong',
    '法规': 'fagui',
    '国际业务': 'guoji',
    '科技创新': 'keji',
    '数字化': 'shuzi',
    '产业金融': 'jinrong',
    '计划与财务': 'jicai',
    '计划财务': 'jicai',
    '计财': 'jicai',
    '输配电': 'shupeidian',
}

ARCH_DIR_MAP = {
    '业务架构覆盖导入日志': 1,
    '数据架构覆盖导入日志': 2,
    '应用架构': 3,
    '网公司应用架构': 3,
}

DROP_SHEET_EXACT = {
    'sheet1', 'Sheet1',
    'WpsReserved_CellImgList',
    '说明',
    '业务能力使能清单',
    '问题统计',
}
DROP_SHEET_PREFIX = ('导出计数_', '临时-')

SKIP_DOMAINS = {'shupeidian', 'jicai'}  # 已有旧版三架构，跳过


def detect_arch_code(zip_dir_name: str) -> int | None:
    for key, code in ARCH_DIR_MAP.items():
        if key in zip_dir_name:
            return code
    return None


def detect_domain_code(filename: str) -> str | None:
    for cn, code in DOMAIN_MAP.items():
        if cn in filename:
            return code
    return None


def clean_excel(src: Path, dst: Path) -> dict:
    """清洗单个 Excel，返回统计信息。"""
    stats = {
        'sheets_removed': [],
        'sheets_kept': [],
        'cols_removed': 0,
    }
    wb = openpyxl.load_workbook(src)

    # 1. 删除无关 sheet
    for sn in list(wb.sheetnames):
        if sn in DROP_SHEET_EXACT or sn.startswith(DROP_SHEET_PREFIX):
            del wb[sn]
            stats['sheets_removed'].append(sn)

    # 2. 删除每个保留 sheet 的最后一列（校验结果列）
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_column > 1:
            ws.delete_cols(ws.max_column, 1)
            stats['cols_removed'] += 1
        stats['sheets_kept'].append(sn)

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    wb.close()
    return stats


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--source-dir', required=True, help='三个解压后的 zip 所在根目录')
    ap.add_argument('--data-dir', default='DATA', help='目标 DATA 目录')
    ap.add_argument('--dry-run', action='store_true', help='只打印不写入')
    args = ap.parse_args()

    src_root = Path(args.source_dir)
    dst_root = Path(args.data_dir)
    if not src_root.exists():
        raise SystemExit(f'source-dir 不存在: {src_root}')

    # 扫描三个架构目录
    planned = []  # [(src_path, dst_path, arch_code, domain_code), ...]
    unmatched = []

    for zip_dir in sorted(src_root.iterdir()):
        if not zip_dir.is_dir():
            continue
        arch_code = detect_arch_code(zip_dir.name)
        if arch_code is None:
            continue

        for xlsx in sorted(zip_dir.glob('*.xlsx')):
            domain_code = detect_domain_code(xlsx.name)
            if domain_code is None:
                unmatched.append(xlsx.name)
                continue
            if domain_code in SKIP_DOMAINS:
                continue
            dst_path = dst_root / domain_code / f'{arch_code}.xlsx'
            planned.append((xlsx, dst_path, arch_code, domain_code))

    # 按域汇总
    by_domain: dict[str, list[int]] = {}
    for _, _, arch, dom in planned:
        by_domain.setdefault(dom, []).append(arch)

    print(f'\n=== 计划处理 {len(planned)} 个 Excel，覆盖 {len(by_domain)} 个新域 ===')
    for dom in sorted(by_domain):
        archs = sorted(by_domain[dom])
        arch_labels = {1: '业务', 2: '数据', 3: '应用'}
        archs_str = '/'.join(arch_labels[a] for a in archs)
        print(f'  {dom:12s}  →  {archs_str}')

    if unmatched:
        print(f'\n=== {len(unmatched)} 个未匹配域名的文件（跳过）===')
        for fn in unmatched:
            print(f'  {fn}')

    skipped = [x for g in ('业务架构', '数据架构', '应用架构', '网公司应用架构')
               for x in src_root.glob(f'*{g}*/*.xlsx')
               if (d := detect_domain_code(x.name)) in SKIP_DOMAINS]
    if skipped:
        print(f'\n=== 跳过已有域（shupeidian/jicai）的 {len(skipped)} 个文件 ===')

    if args.dry_run:
        print('\n[DRY-RUN] 未写入任何文件。')
        return

    # 实际执行
    print(f'\n=== 开始清洗并写入 {dst_root} ===')
    total_cols = 0
    total_sheets_removed = 0
    for i, (src, dst, _, dom) in enumerate(planned, 1):
        stats = clean_excel(src, dst)
        total_cols += stats['cols_removed']
        total_sheets_removed += len(stats['sheets_removed'])
        print(f'  [{i:3d}/{len(planned)}] {dom}/{dst.name}  '
              f'删列={stats["cols_removed"]}  '
              f'删sheet={len(stats["sheets_removed"])}  '
              f'保留sheet={len(stats["sheets_kept"])}')

    print(f'\n=== 完成 ===')
    print(f'  Excel 文件:  {len(planned)}')
    print(f'  域目录:      {len(by_domain)}')
    print(f'  总删列数:    {total_cols}')
    print(f'  总删sheet数: {total_sheets_removed}')


if __name__ == '__main__':
    main()
